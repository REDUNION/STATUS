# -*- coding: utf-8 -*-

"""
Module implementing status.
"""
from __future__ import print_function
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QMessageBox
from PyQt5 import QtGui
from PyQt5 import QtWidgets
from .Ui_status import Ui_status
from openpyxl import * 
from openpyxl.styles import *                 
#from openpyxl.styles.colors import *
from datetime import datetime
import time 


cur = datetime.now()
y = cur.year
m = cur.month
d = cur.day
hour =7

imerom = cur.strftime("%d/%m/%y")
clock = cur.strftime("%H:%M")

class status(QMainWindow, Ui_status):
    """
    Class documentation goes here.
    """
    def __init__(self, parent=None):
        """
        Constructor
        
        @param parent reference to the parent widget
        @type QWidget
        """
        super(status, self).__init__(parent)
        self.setupUi(self)
        self.txt_apo.setText("07:00")
        self.start = datetime(year=y,
                        month=m,
                        day=d,
                        hour=7,
                        minute=00,
                        second=00)

        d1_ts = time.mktime(self.start.timetuple())
        d2_ts = time.mktime(cur.timetuple())
        self.diarkia=str(int((d2_ts-d1_ts)/60))

    @pyqtSlot()
    def on_addrow_clicked(self):
        """
        ΦΟΡΤΩΣΗ  ΤΩΝ ΤΙΜΩΝ ΣΤΟ topiko  EXCEL
        """
        numRows = self.excel.rowCount()
        self.excel.insertRow(numRows)
        self.excel.setItem(numRows, 0, QtWidgets.QTableWidgetItem(self.txt_onoma.text()))
        self.excel.setItem(numRows, 1, QtWidgets.QTableWidgetItem(self.txt_work.text()))
        self.excel.setItem(numRows, 2, QtWidgets.QTableWidgetItem(self.txt_posot.text()))
        self.excel.setItem(numRows, 3, QtWidgets.QTableWidgetItem(self.shmera_2.text()))
        self.excel.setItem(numRows, 4, QtWidgets.QTableWidgetItem(self.txt_apo.text()))
        self.excel.setItem(numRows, 5, QtWidgets.QTableWidgetItem(self.txt_ews.text()))  
        self.excel.setItem(numRows, 6, QtWidgets.QTableWidgetItem(self.txt_parat.text()))
        

        self.txt_apo.setText(time.strftime("%H:%M"))
        self.start = time.strftime("%H:%M")

    @pyqtSlot()
    def on_preview_clicked(self):
        """
        ΠΡΟΕΠΙΣΚΟΠΗΣΗ 
        """
#        start = self.txt_apo.text()
        self.txt_ews.setText(time.strftime("%H:%M"))
        self.diarkeia.setText(self.diarkia)
        self.shmera_2.setText(cur.strftime("%d/%m/%Y"))

    @pyqtSlot()
    def on_submit_excel_clicked(self):
        """
        Slot documentation goes here.
        """   
  

        wb = load_workbook(filename = self.filepath )
        ws = wb.active    
        yrow=self.excel.rowCount()
        ws.insert_rows(idx=2,amount=yrow) 
        for currentColumn in range(1, self.excel.columnCount()+1):
              for currentRow in range(1, self.excel.rowCount()+1):
                    try:
                         teext = str(self.excel.item(currentRow-1, currentColumn-1).text())
                         ws.cell(currentRow+1, currentColumn).protection = Protection(locked = False ,  hidden = False )
                         ws.cell(currentRow+1, currentColumn).value = teext
                         ws.cell(currentRow+1, currentColumn).font = Font(name='Arial', bold=True, size=13)
                         ws.cell(currentRow+1, currentColumn).alignment = Alignment(horizontal='center',vertical='center', wrap_text= True) 
                         ws.cell(currentRow+1, currentColumn).border = Border(left=Side(border_style='thick',color='FF000000'),right=Side(border_style='thick',color='FF000000'),top=Side(border_style='double',color='FF000000'),bottom=Side(border_style='double',color='FF000000'))

                    except AttributeError:
                          pass

        QMessageBox.about(self, "ΗΜΕΡΙΣΙΑ ΚΑΤΑΚΩΧΗΣΗ STATUS", "ΠΡΟΚΕΙΤΑΙ ΝΑ ΚΑΤΑΧΩΡΥΘΟΥΝ ΕΙΣΑΣΤΕ ΣΥΜΦΩΝΟΣ ")
        wb.save(filename = self.filepath)   
        QMessageBox.about(self, "ΗΜΕΡΙΣΙΑ ΚΑΤΑΚΩΧΗΣΗ STATUS", "ΟΚ ΚΑΤΑΧΩΡΗΘΗΚΑΝ")          





    @pyqtSlot()
    def on_browse_xl_clicked(self):
        """
        ΒΡΕΣ ΤΟ ΑΡΧΕΙΟ ΠΟΥ ΕΙΝΑΙ ΠΡΟΣ ΕΠΕΞΕΡΓΑΣΙΑ
        """
       
        self.filepath = QFileDialog.getOpenFileName(self, "ΑΝΟΙΞΕ ΤΟ", "ΑΝΟΙΞΕ ΤΟ", "STATUS (*.xlsx)")[0]
        print(self.filepath)

        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.xl_directory.setFont(font)

        self.xl_directory.setText(self.filepath)
    
    @pyqtSlot()
    def on_auto_2_clicked(self):
        """
        Slot documentation goes here.
        """
   

    @pyqtSlot()
    def on_deleterow_clicked(self):
        """
        σβηνει ΤΗΝ ΕΠiΛΕΓΜεΝΗ ΣΕΙΡΑ
        """
        selected = self.excel.currentIndex()
        self.excel.removeRow(selected.row())
    
    @pyqtSlot()
    def on_clear_clicked(self):
        """
        ΚΑΘΑΡΙΣΜΟΣ ΤΟΥ EXCEL
        """
        self.excel.setRowCount(0)
    
    @pyqtSlot()
    def on_about_clicked(self):
        """
        ABOUT ME
        """
        QMessageBox.about(self, "copyright STATUS_A_TMHMATOS.VC1 alkis amanatidis", "creator αμανατιδης αλκης \nemail:amanatidisalkis@gmail.com ")

